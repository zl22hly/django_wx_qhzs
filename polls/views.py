# import os
# from django.http import HttpResponse
# def index(request):
#     return HttpResponse("Hello, world. You're at the polls .")
# def test(request):
#     return HttpResponse("Hello.")
# -------------------------
# from rest_framework import serializers
# from polls.models import park
# from rest_framework import serializers
# ++++++++++
from django.shortcuts import render, redirect
from rest_framework.viewsets import ModelViewSet
from rest_framework.pagination import LimitOffsetPagination
from rest_framework import filters
from .serializers import ParkInfoSerializer, HomeDataInfoSerializer, ProtectInfoSerializer, UserInfoSerializer, \
    NewsInfoSerializer, NewsPostsInfoSerializer, HomeIMGDataInfoSerializer
from .models import HomeData, Park, Protect, User, News, NewsPosts, HomeIMGData
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import SearchFilter, OrderingFilter
from django import forms
from openpyxl import load_workbook


import os
import sys


class UserView(ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserInfoSerializer


class HomeView(ModelViewSet):
    queryset = HomeData.objects.all()
    serializer_class = HomeDataInfoSerializer


class NewsView(ModelViewSet):
    queryset = News.objects.all()
    serializer_class = NewsInfoSerializer
    pagination_class = LimitOffsetPagination
    filter_backends = (SearchFilter, OrderingFilter,
                       DjangoFilterBackend)  # 指定过滤器
    search_fields = ('category', 'id')  # 指定可搜索的字段
    filterset_fields = ('category','id' )
    filter_backends = [filters.OrderingFilter]
    ordering_fields=['id','add_Date']
# /news/?category=1&ordering=-id&limit=20&offset=0


class NewsPostsView(ModelViewSet):
    queryset = NewsPosts.objects.all()
    serializer_class = NewsPostsInfoSerializer


class HomeIMGDataView(ModelViewSet):
    queryset = HomeIMGData.objects.all()
    serializer_class = HomeIMGDataInfoSerializer


class ParkView(ModelViewSet):
    queryset = Park.objects.all()
    serializer_class = ParkInfoSerializer
    filter_backends = (SearchFilter, OrderingFilter,
                       DjangoFilterBackend)  # 指定过滤器
    search_fields = ('name1', 'fullname', 'id')  # 指定可搜索的字段
    filterset_fields = ('name1', 'fullname', 'id')


class ProtectView(ModelViewSet):
    queryset = Protect.objects.all()
    serializer_class = ProtectInfoSerializer
    # filter_backends = (SearchFilter, OrderingFilter, DjangoFilterBackend)  # 指定过滤器
    # search_fields = ('name2', )  # 指定可搜索的字段
    # filterset_fields = ('name2', )


class parkFormAdd(forms.ModelForm):
    class Meta:
        model = Park
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for name, field in self.fields.items():
            field.widget.attrs = {"class": "form-control"}


def home_list(request):
    queryset = User.objects.all()
    return render(request, 'home.html')


def park_list(request):
    queryset = Park.objects.all()
    return render(request, 'park.html', {'queryset': queryset})


def park_add(request):
    if request.method == "GET":
        form = parkFormAdd()
        return render(request, 'parkadd.html', {'form': form})
    form = parkFormAdd(data=request.POST)
    if form.is_valid():
        # print(form.cleaned_data)
        form.save()
        return redirect("/parklist/")
    else:
        return render(request, 'parkadd.html', {'form': form})


def park_delete(request):
    nid = request.GET.get("nid")
    Park.objects.filter(id=nid).delete()
    return redirect("/parklist/")


def park_edit(request, nid):
    row_object = Park.objects.filter(id=nid).first()


def add_park(request):
    if request.method == "GET":
        return render(request, 'addPark.html')
    file_obj = request.FILES.get("XLS")
    wd = load_workbook(file_obj)
    print(type(wd))
    sheet = wd.worksheets[0]
    for row in sheet.iter_rows(min_row=2):
        naturalLevel_v = row[0].value
        ParkLevel_v = row[1].value
        numbering_v = row[2].value
        name1_v = row[3].value
        fullname_v = row[4].value
        abbreviation_v = row[5].value
        icon_v = row[6].value
        initiate_v = row[7].value
        slide_v = row[8].value
        introduceEN_v = row[9].value
        introduceZH_v = row[10].value
        introduceMP3EN_v = row[11].value
        introduceMP3ZH_v = row[12].value
        introduceMP4_v = row[13].value
        introduceVR_v = row[14].value
        lon_v = row[15].value
        lat_v = row[16].value
        mapLevel_v = row[17].value
        province_v = row[18].value
        area_v = row[19].value
        city_v = row[20].value
        county_v = row[21].value
        scenery_v = row[22].value
        exists = Park.objects.filter(name1=name1_v).exists()
        if not exists:
            Park.objects.create(naturalLevel=naturalLevel_v, ParkLevel=ParkLevel_v,
                                numbering=numbering_v, name1=name1_v, fullname=fullname_v,
                                abbreviation=abbreviation_v, icon=icon_v,
                                initiate=initiate_v, slide=slide_v,
                                introduceEN=introduceEN_v, introduceZH=introduceZH_v,
                                introduceMP3EN=introduceMP3EN_v, introduceMP3ZH=introduceMP3ZH_v,
                                introduceMP4=introduceMP4_v, introduceVR=introduceVR_v,
                                lon=lon_v, lat=lat_v,
                                mapLevel=mapLevel_v, province=province_v,
                                area=area_v, city=city_v,
                                county=county_v, scenery=scenery_v)
        else:
            Park.objects.filter(name1=name1_v).update(naturalLevel=naturalLevel_v, ParkLevel=ParkLevel_v,
                                                      numbering=numbering_v, name1=name1_v, fullname=fullname_v,
                                                      abbreviation=abbreviation_v, icon=icon_v,
                                                      initiate=initiate_v, slide=slide_v,
                                                      introduceEN=introduceEN_v, introduceZH=introduceZH_v,
                                                      introduceMP3EN=introduceMP3EN_v, introduceMP3ZH=introduceMP3ZH_v,
                                                      introduceMP4=introduceMP4_v, introduceVR=introduceVR_v,
                                                      lon=lon_v, lat=lat_v,
                                                      mapLevel=mapLevel_v, province=province_v,
                                                      area=area_v, city=city_v,
                                                      county=county_v, scenery=scenery_v)

    return render(request, 'addPark.html')


def add_news(request):
    # 返回基础界面
    if request.method == "GET":
        return render(request, 'addNews.html')
    # 接受上传的对象
    file_obj = request.FILES.get("XLS")
    # 载入插件
    wd = load_workbook(file_obj)
    print(type(wd))
    # 第一个表格
    sheet = wd.worksheets[0]
    # 从第二行开始
    for row in sheet.iter_rows(min_row=2):
        fatherParkID= NewsPosts.objects.get(id=row[8].value)
        fatherPark_v = fatherParkID
        IMG_v = row[1].value
        title_v = row[2].value
        subTitle_v = row[3].value
        point_v = row[4].value
        text_v = row[5].value
        quote_v = row[6].value
        add_Date_v = row[7].value
        category_v = row[8].value

        # 数据库是否存在name1的
        exists = News.objects.filter(title=title_v).exists()
        # 不存在就新建存在就修改
        if not exists:
            News.objects.create(fatherPark=fatherPark_v,
                                IMG=IMG_v,
                                title=title_v,
                                subTitle=subTitle_v,
                                point=point_v,
                                text=text_v,
                                quote=quote_v,
                                add_Date=add_Date_v,
                                category=category_v,
                                )
        else:
            News.objects.filter(title=title_v).update(fatherPark=fatherPark_v,
                                                      IMG=IMG_v,
                                                      title=title_v,
                                                      subTitle=subTitle_v,
                                                      point=point_v,
                                                      text=text_v,
                                                      quote=quote_v,
                                                      add_Date=add_Date_v,
                                                      category=category_v,
                                                      )

    return render(request, 'addNews.html')
